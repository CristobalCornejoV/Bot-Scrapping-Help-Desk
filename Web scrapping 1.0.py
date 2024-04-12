from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import time
import re
from openpyxl import Workbook, load_workbook

# Opciones de configuración para el navegador Firefox
options = Options()
options.binary_location = "ruta_al_ejecutable_de_firefox"

# Crear una instancia del navegador Firefox
driver = webdriver.Firefox(options=options)

# URL de inicio de sesión (reemplaza con tu URL)
url = "URL_del_sitio_web"

# Abrir la página de inicio de sesión
driver.get(url)

# Esperar hasta 10 segundos para que se cargue la página
driver.implicitly_wait(10)

# Iniciar sesión con la cuenta de UdeChile
boton_iniciar_udechile = driver.find_element(By.XPATH, "//span[contains(text(),'Iniciar con UdeChile')]")
boton_iniciar_udechile.click()

driver.implicitly_wait(10)

# Ingresar nombre de usuario y contraseña (reemplaza con tus credenciales)
campo_usuario = driver.find_element(By.ID, 'usernameInput')
campo_usuario.send_keys("usuario")

campo_password = driver.find_element(By.ID, 'passwordInput')
campo_password.send_keys("contraseña")

# Hacer clic en el botón de ingresar
boton_ingresar = driver.find_element(By.XPATH, "//button[contains(text(),'Ingresar')]")
boton_ingresar.click()

# Esperar 10 segundos después de iniciar sesión
driver.implicitly_wait(10)
time.sleep(5)

# Crear un nuevo archivo Excel y agregar encabezados
wb_destino = Workbook()
ws_destino = wb_destino.active
ws_destino.append(["Inventario Encontrado", "Correo solicitante"])

# Leer el archivo Excel y obtener los IDs (reemplaza con la ruta de tu archivo Excel)
workbook = load_workbook(filename="ruta_al_archivo_excel_original")
sheet = workbook.active
columna_a = sheet['A']

for cell in columna_a:
    id = cell.value
    if id: 
        url_ticket = f"URL_del_ticket_con_id_{id}"
        driver.get(url_ticket)
        time.sleep(3)

        # Encontrar el número de inventario del PC dentro del elemento con id "frmCampos:dtHistoryItems_data"
        try:
            elemento_historial = driver.find_element(By.ID, "frmCampos:dtHistoryItems_data")
            lineas = elemento_historial.text.split('\n')
            numero_inventario = "Código no encontrado"
            for linea in lineas:
                match_pc = re.search(r'Número\s*de\s*inventario\s*del\s*PC\s*:\s*(\S+)', linea)
                match_igeo = re.search(r'Número\s*de\s*inventario\s*\(\s*IGEO\s*\)\s*del\s*computador\s*:\s*(\S+)', linea)
                if match_pc:
                    numero_inventario = match_pc.group(1)
                    break
                elif match_igeo:
                    numero_inventario = match_igeo.group(1)
                    break

            # Encontrar el título dentro del elemento con la clase "profile" y dentro de "cuttedText"
            elemento_profile = driver.find_element(By.CLASS_NAME, "profile")
            elemento_cutted_text = elemento_profile.find_element(By.CLASS_NAME, "cuttedText")
            title = elemento_cutted_text.text
            
            # Agregar los datos al archivo Excel destino
            ws_destino.append([numero_inventario, title])

            # Guardar el archivo Excel destino
            wb_destino.save("ruta_al_archivo_excel_destino")

            # Presionar el botón de salir para evitar bloqueos
            wait = WebDriverWait(driver, 10)
            boton_salir = wait.until(EC.visibility_of_element_located((By.XPATH, "//ul[@id='ulBar']//div[@id='frmTopBar:pnlGrpTopBar']//button[@id='frmTopBar:closeButton']")))
            boton_salir.click()
            
        except Exception as e:
            print("Error:", e)

# Cerrar el navegador después de terminar
driver.quit()
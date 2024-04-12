from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import time
import re

# Rutas de archivos y ubicaciones
ruta_firefox = "ruta_al_ejecutable_de_firefox"
ruta_excel = "ruta_al_archivo_excel_original"
ruta_excel_destino = "ruta_al_archivo_excel_destino"

# Configuración de opciones de Firefox
options = Options()
options.binary_location = ruta_firefox

# Inicializar el navegador Firefox
driver = webdriver.Firefox(options=options)

# URL de inicio de sesión
url = "URL_del_sitio_web"
driver.get(url)

# Esperar implícitamente hasta 10 segundos para encontrar elementos
driver.implicitly_wait(10)

# Iniciar sesión con UdeChile
boton_iniciar_udechile = driver.find_element(By.XPATH, "//span[contains(text(),'Iniciar con UdeChile')]")
boton_iniciar_udechile.click()

# Esperar implícitamente hasta 10 segundos para encontrar elementos
driver.implicitly_wait(10)

# Ingresar credenciales de usuario
campo_usuario = driver.find_element(By.ID, 'usernameInput')
campo_usuario.send_keys("usuario")

campo_password = driver.find_element(By.ID, 'passwordInput')
campo_password.send_keys("contraseña")

# Hacer clic en el botón de ingresar
boton_ingresar = driver.find_element(By.XPATH, "//button[contains(text(),'Ingresar')]")
boton_ingresar.click()

# Esperar implícitamente hasta 10 segundos para encontrar elementos
driver.implicitly_wait(10)
time.sleep(5)

# Leer el archivo Excel y obtener los IDs
workbook = load_workbook(filename=ruta_excel)
sheet = workbook.active
columna_a = sheet['A']

# Contar la cantidad de tickets en el archivo Excel
cantidad_tickets = sum(1 for cell in columna_a if cell.value)

# Preguntar al usuario si desea leer los tickets
respuesta = input(f"¿Quieres leer {cantidad_tickets} tickets? (y/n): ")

if respuesta.lower() == 'y':
    # Crear un nuevo archivo Excel y agregar encabezados
    wb_destino = Workbook()
    ws_destino = wb_destino.active
    ws_destino.append(["Inventario Encontrado", "Correo solicitante"])

    for cell in columna_a:
        id = cell.value
        if id:
            # Construir la URL del ticket
            url_ticket = f"URL_del_ticket_con_id_{id}"
            driver.get(url_ticket)
            time.sleep(3)

            try:
                # Encontrar el elemento de historial
                elemento_historial = driver.find_element(By.ID, "frmCampos:dtHistoryItems_data")
                lineas = elemento_historial.text.split('\n')
                numero_inventario = "Código no encontrado"
                for linea in lineas:
                    # Buscar número de inventario del PC
                    match_pc = re.search(r'Número\s*de\s*inventario\s*del\s*PC\s*:\s*(\S+)', linea)
                    match_igeo = re.search(r'Número\s*de\s*inventario\s*\(\s*IGEO\s*\)\s*del\s*computador\s*:\s*(\S+)', linea)
                    if match_pc:
                        numero_inventario = match_pc.group(1)
                        break
                    elif match_igeo:
                        numero_inventario = match_igeo.group(1)
                        break

                # Encontrar el título
                elemento_profile = driver.find_element(By.CLASS_NAME, "profile")
                elemento_cutted_text = elemento_profile.find_element(By.CLASS_NAME, "cuttedText")
                title = elemento_cutted_text.text

                # Imprimir los datos obtenidos
                print("Número de inventario del PC:", numero_inventario)
                print("Título dentro en profile:", title)

                # Agregar los datos al archivo Excel destino
                ws_destino.append([numero_inventario, title])

                # Guardar el archivo Excel destino
                wb_destino.save(ruta_excel_destino)

                # Presionar salir para no tener candado
                wait = WebDriverWait(driver, 10)
                boton_salir = wait.until(EC.visibility_of_element_located((By.XPATH, "//ul[@id='ulBar']//div[@id='frmTopBar:pnlGrpTopBar']//button[@id='frmTopBar:closeButton']")))
                boton_salir.click()

            except Exception as e:
                print("Error:", e)

    print("Se han leído los tickets y guardado en el archivo Excel destino.")
else:
    print("No se leyeron los tickets. Saliendo del programa.")

# Cerrar el navegador después de terminar
driver.quit()